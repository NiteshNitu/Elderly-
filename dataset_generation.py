import openpyxl
import time
import requests

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Sheet 1"
headings = ['Ax' , 'Ay' , 'Az' , 'Gx' , 'Gy' , 'Gz']
for i in range(1,7):
    c = sheet.cell(1 , i)
    c.value = headings[i-1]
for _ in range(4):
    for i in range(1,501):
        response = requests.get('http://192.168.1.106/Python')
        res = response.text.split(sep = " ")
        a = []
        max = sheet.max_row
        for j in range(1,7):
            a.append(float(res[j-1]))
            c = sheet.cell(max + 1 , j)
            c.value = a[j-1]
        print(a , end = "  ")
        print(i)
        wb.save("afterg1{}.xlsx".format(_+5))
        a.clear()
        time.sleep(0.65)